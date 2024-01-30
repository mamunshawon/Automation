import time

import openpyxl
import requests
from pandas import read_excel
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait

# Define your bot token and chat ID
bot_token = '6510460079:AAEzl9SdC2yKpHCPFfs4f0-Een8k02H3FTc'
chat_id = '-972000340'


def send_telegram_message(bot_token, chat_id, message):
    url = f'https://api.telegram.org/bot{bot_token}/sendMessage'
    data = {
        'chat_id': chat_id,
        'text': message
    }
    response = requests.post(url, data = data)
    return response


# Open the Excel file and read the data into a pandas dataframe
data = read_excel("MNOCAM.xlsx")
username = "uatdemo18@gmail.com"
password = "N@gad1234"

driver = webdriver.Chrome()
driver.get('https://systest.mynagad.com:20020/ui/system/#/home')
driver.maximize_window()
time.sleep(3)

driver.find_element(By.ID, "username").send_keys(username)
driver.find_element(By.ID, "password").send_keys(password)

try:
    driver.find_element(By.ID, "login_button").click()
except:
    print("login failed.......")
    time.sleep(2)

# Extract row names from the DataFrame
row_names = data.iloc[:, 0].tolist()

# Initialize an empty report string
full_report = ""

for index, row in data.iterrows():
    driver.get('https://systest.mynagad.com:20020/ui/system/#/campaign/list')
    time.sleep(3)
    Enter_Campaign_Name = driver.find_element(by = By.XPATH, value = "/html/body/app-root/app-full-layout/div/div["
                                                                     "2]/div/div/div/app-campaign-list/section/div["
                                                                     "2]/div/div[1]/div/ngb-accordion/div/div["
                                                                     "2]/div/div/form/div/div[1]/div[1]/div/input")
    Enter_Campaign_Name.send_keys(row["Name"])
    time.sleep(2)
    Click_Search_button = driver.find_element(by = By.XPATH, value = '/html/body/app-root/app-full-layout/div/div['
                                                                     '2]/div/div/div/app-campaign-list/section/div['
                                                                     '2]/div/div[1]/div/ngb-accordion/div/div['
                                                                     '2]/div/div/form/div/div[2]/div/button')
    Click_Search_button.click()
    time.sleep(2)
    # Define the text you want to search for
    search_text = "TEST MNO"

    # Wait for the page to load and find the element containing the text
    try:
        # Use an explicit wait for the element to be present
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{search_text}')]"))
        )

        # Extract and print the text content of the located element
        text_content = element.text
        print(f"MATCH CAMPAIGN NAME: {text_content}")
        # Find the button by link text and click on it
        # Replace with the actual link text of your button
        Click_Details_button = driver.find_element(by = By.XPATH, value = "/html/body/app-root/app-full-layout/div/div["
                                                                          "2]/div/div/div/app-campaign-list/section"
                                                                          "/div["
                                                                          "2]/div/div[2]/div/div["
                                                                          "2]/div/app-common-table-advanced/table"
                                                                          "/tbody/tr["
                                                                          "1]/td[5]/div/span[1]/button")
        Click_Details_button.click()
        time.sleep(2)
        # Read search terms from Excel
        excel_file_path = "C:/Users/mamunur.shawon/PycharmProjects/testpro/MNOCAM.xlsx"  # Replace with the actual
        # path to
        # your Excel file
        excel_sheet_name = "Sheet1"  # Replace with the actual sheet name

        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[excel_sheet_name]

        # Assuming your data starts from the second row and second column
        for row in range(2, sheet.max_row + 1):
            for column in range(2, sheet.max_column + 1):
                # Convert the value to string
                search_text = str(sheet.cell(row = row, column = column).value)
                # Wait for the page to load and find the element containing the text
                try:
                    element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{search_text}')]"))
                    )
                    # Extract and print the text content of the located element
                    text_content = element.text
                    print(f"Found text: {text_content} for search term: {search_text}")
                except Exception as e:
                    print(f"Error: {e} for search term: {search_text}")

        Click_show_Details = driver.find_element(by = By.XPATH, value = "/html/body/app-root/app-full-layout/div/div["
                                                                        "2]/div/div/div/app-campaign-details/section"
                                                                        "/div[2]/div/div["
                                                                        "2]/div/div/div/form/div/div/div["
                                                                        "1]/div/table/tbody/tr[1]/td[2]/button[1]")
        Click_show_Details.click()
        time.sleep(3)
        # Read search terms from Excel
        excel_file_path = "C:/Users/mamunur.shawon/PycharmProjects/testpro/MNOCAM.xlsx"
        # Replace with the actual path to your Excel file
        excel_sheet_name = "Sheet2"  # Replace with the actual sheet name
        # Open the Excel file
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook[excel_sheet_name]
        # Assuming your data starts from the second row and second column
        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                # Convert the value to string
                search_text = str(sheet.cell(row = row, column = column).value)
                # Wait for the page to load and find the element containing the text
                try:
                    element = WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{search_text}')]"))
                    )
                    # Extract and print the text content of the located element
                    text_content = element.text
                    print(f"Found text: {text_content} for search term: {search_text}")
                except Exception as e:
                    print(f"Error: {e} for search term: {search_text}")
    except Exception as e:
        print(f"{e}")
driver.close()
